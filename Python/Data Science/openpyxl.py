""" Excel files """
from openpyxl import Workbook
import datetime

#Create a worksheet
wb = Workbook()    #creates workbook object
sheet = wb.active  #creates sheet 

ws = Wb.create_sheet(   #creates new sheet
    "Sheet name",      #name of the sheet
    x)                 #create new sheet
ws.title = "New Title" #change title of sheet
wks = wb["New Title"]  #save key of workbook

#Print names of worksheets
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)
    
#copy worksheets in a workbook
source = wb.active
target = wb.copy_worksheet(source)

        """ Modifying & Accessing Data """   
#Getting data
ws["A"]     #Get all cell names in column A
ws["A:B"]   #Range from column A to B
ws["A1":'C2']  #all cells from certain range
ws[x]       #Get all cells from row x
ws[x:y]     #Range from rowx to rowy

.value        #Returns value of a cell.
.title        #Reads title of Sheet         
ws["A1"].value  #Returns value of A1 cell
.cell(row = x, column = y)  #which cell it is
.cell(x,y).value     #value of cell xy


#Writing in data
ws["A1"] = "hello"   #assign value to cell
ws['B1'] = datetime.datetime(2010,7,21) #write date
   ws['B1'].number_format
ws['C1'] = "SUM(1,1)"   #write forumula
a1 = ws["A1"]
a1.value = "value"

ws.merge_cells('A2:D2')  #merges all cells from top-left
ws.unmerge_cells('A2:D2') #unmerges cells
ws.merge_cells(start_row=x, start_column=y, end ~~~)

data = [        #How to add in entire columns of data
    ["col1", "col2"],
    ["r1c1", "r1c2"]]

for r in data:       #append in a loop
    ws.append(r)

ws.auto_filter.ref = "A1:B15"
ws.auto_filter.add_filter_column(0, ["Kiwi", "Apple", "Mango"])
ws.auto_filter.add_sort_condition("B2:B15")

#Managing rows and columns
.insert_rows(       #inserts rows
.delete_rows(       #deletes'rows
.insert_cols(       #inserts columns
.delete_cols(       #deletes columns
    .idx = x,        #inserts before x, deletes at x
    .amount = y)     #how many are added/removed
.move_range("D4:F10", rows=-1, cols=2)


#Iterate through entire datasheet at once
tuple(ws.rows)      #prints the rows names
tuple(ws.columns)

for row in ws.values:   #get all the values
    for value in row:
        print(value)

#printing out column headings
for value in sheet.iter_rows(
        min_row=1,
        max_row=1,
        values_only=True):
    print(value)


#Tables
#create a table

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()
ws = wb.active

data = [
    ['Apples', 10000, 5000, 8000, 6000],
    ['Pears',   2000, 3000, 4000, 5000],
    ['Bananas', 6000, 6000, 6500, 6000],
    ['Oranges',  500,  300,  200,  700]]

# add column headings. these must be strings
ws.append(["Fruit", "2011", "2012", "2013", "2014"])
for row in data:
    ws.append(row)

tab = Table(displayName="Table1", ref="A1:E5")

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style

ws.add_table(tab)
wb.save("table.xlsx")

# Get table by name or range
ws.tables['Table name']  #name
ws.tables["A1:D10"]      #range
ws.tables.items()     #list of table name and ranges


del ws.tables["Table"]  #delete a table
len(ws.tables)    #get number of tables in worksheet

for table in ws.tables.values():
    print(table)


#Working with Pandas Dataframes
from openpyxl.utils.dataframe import dataframe_to_rows
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)
    
for cell in ws["A"] + ws[1]:
    cell.style = 'Pandas'

#Insert an image
from openpyxl.drawing.image import Image
img = Iamge("image_file.png")
ws.add_image(img, 'A1')

#Adding a comment to a cell
from openpyxl.comments import Comment
comment = ws["A1"].comment
comment = Comment("Comment text", "comment Author")
  .comment = #assigns comment
comment.text    #prints out the comment text
comment.author   #prints out the author of the comment
comment.width = x   #how wide the comment is
comment.height = y  #how tall the comment is

#Change font
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
font = Font(
    name = '',  #font style
    size = x,     #size of font
    bold = ,    #boolean, True or False
    italic = ,   #boolean, True or False
    vertAlign = ,   #alignment
    underline = ''   #underline
    strike = ,  #boolean. strikethrough
    color = '')  #color of text
fill = PatternFill(
    fill_type = None,   #The fill color
    start_color = "",   #The start color of fill
    end_color = "",)     #ending color of fill

border = Border(    #border of the box
    left, right, top, botton, 
    diagonal, outline, 
    vertical, horizontal=Side(
        border_style = ,  #style of border
        color = ""))      #color of the border

from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
a1 = ws['A1']
ft = Font(color="FF0000")
a1.font = ft
a1.font.italic = True # is not allowed # doctest: +SKIP
a1.font = Font(color="FF0000", italic=True) # the change only affects A1

a2 = copy(a1)  #copy a1
a2.name = "name"

            """ Charts & graphs """
for i in range(10):
    ws.append([i])

from openpyxl.chart import BarChart, Reference, Series
values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
chart = BarChart()
chart.add_data(values)
ws.add_chart(chart, "E15")

           #Area charts
#2D Area Chars
from openpyxl.chart import (AreaChart, Reference, Series)

wb = Workbook()
ws = wb.active

rows = [['Number', 'Batch 1', 'Batch 2'], #column headings
    [2, 40, 30],   #Row 2
    [3, 40, 25]]

for row in rows:
    ws.append(row)

chart = AreaChart()  #Create chart. AreaChart3D() for 3D
chart.title = "Area Chart"  #title of chart
chart.style = 13    #
chart.x_axis.title = 'Test'  #title of x-axis
chart.y_axis.title = 'Percentage'  #title of y-axis
chart.legend = None    #for the legend

cats = Reference(ws, min_col=1, min_row=1, max_row=7)
data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws.add_chart(chart, "A10")

           #Bar Charts
from openpyxl.chart import BarChart, Series, Reference

wb = Workbook(write_only=True)
ws = wb.create_sheet()
rows = [('Number', 'Batch 1', 'Batch 2'),
    (2, 10, 30)]

for row in rows:
    ws.append(row)

chart1 = BarChart()   #creates bar chart
.type =         #How bars are oriented
     "col",       #Vertical bars
     "bar"        #Horizontal bars
chart1.style =  #Style of chart.
     10,          #standard bar chart
     11,          #Horizontal chart
     12,          #stacked chart
     13           #stacked horizontal
.grouping =     #How bars are laid out
     'stacked'        #two bars on top of each other
     'percentStacked' #percent bars on top
.title = "Bar Chart"  #title
.y_axis.title = 'y-axis label'
.x_axis.title = 'x-axis label'

data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
cats = Reference(ws, min_col=1, min_row=2, max_row=7)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws.add_chart(chart1, "Cell")  #Top-left cell where graph starts
       
            #Bubble Chart
from openpyxl.chart import Series, Reference, BubbleChart

rows = [
    ("Number of Products", "Sales in USD", "Market share"),
    (14, 12200, 15),
    (20, 60000, 33)]

for row in rows:
    ws.append(row)

chart = BubbleChart()
chart.style = 18 # use a preset style

# add the first series of data
xvalues = Reference(ws, min_col=1, min_row=2, max_row=5)
yvalues = Reference(ws, min_col=2, min_row=2, max_row=5)
size = Reference(ws, min_col=3, min_row=2, max_row=5)
series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title="2013")
chart.series.append(series)

# add the second
xvalues = Reference(ws, min_col=1, min_row=7, max_row=10)
yvalues = Reference(ws, min_col=2, min_row=7, max_row=10)
size = Reference(ws, min_col=3, min_row=7, max_row=10)
series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title="2014")
chart.series.append(series)

# place the chart starting in cell E1
ws.add_chart(chart, "E1")
wb.save("bubble.xlsx")


        """ Loading & saving workbook """
#Saving to a file
save_path = "folder to save file in"
wb = Workbook()
wb.save(save_path + "file_name.xlsx")

#Saving as a stream
from tempfile import NamedTemporaryFile
with NamedTemporaryFile() as tmp:
    wb.save(tmp.name)
    tmp.seek(0)
    stream = tmp.read

#Loading from a file
from openpyxl import load_workbook
workbook = load_workbook(
    'file_to_read',
    #read_only=True, #read-only mode
    #write_only=True 
    #data_only) #ignores formula
workbook.sheetnames   #prints out names of sheets

#Save worksheet as a dataframe
df = DataFrame(ws.values)

#If it has headers or indices 
from itertools import islice
data = ws.values
cols = next(data)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 1, None) for r in data)
df = DataFrame(data, index=idx, columns=cols)

#save information into a dictionary
import json
dicts = {}
for row in sheet.iter_rows(min_row=2,
                           min_col=4,
                           max_col=7,
                           values_only=True):
    product_id = row[0]
    product = {
        "column1": row[1],
        "column2": row[2],
        "column3": row[3]
    }
    products[product_id] = product
print(json.dumps(products))   #save for later display






