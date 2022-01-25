# Duracell code. Copy data from 30+ files into one file to represent monthly data.

# importing openpyxl module
import openpyxl as xl;
  
# opening the source excel file. LOOP THIS
source_folder = """Source folder here"""
destination_file = "file name"

# Looping through each file in source folder
for file in source_folder:
  filename = file
  wb1 = xl.load_workbook(filename)
  ws1 = wb1.worksheets[0]
  
  # opening the destination excel file. NEED THIS OPEN IN LOOP TO GET CORRECT ROW & COLUMN NUMBER.
  filename1 = destination_file
  wb2 = xl.load_workbook(filename1)
  ws2 = wb2.active

  # calculate total number of rows and 
  # columns in source excel file
  mr1 = ws1.max_row
  mc1 = ws1.max_column
  
  mr2 = ws2.max_row  # reads max rows and columns of destination file so you 
  mc2 = ws2.max_column # don't constantly copy over it.
  
  # copying the cell values from source 
  # excel file to destination excel file
  for i in range (2, mr1 + 1):  #Copies from row 2 to last row
      for j in range (1, mc1 + 1): #Copes from column 1 to last column
          # reading cell value from source excel file
          c = ws1.cell(row = i, column = j)

          # writing the read value to destination excel file
          ws2.cell(row = mr2 + i, column = mr2 + j).value = c.value

  # saving the destination excel file
  wb2.save(str(filename1))
  
