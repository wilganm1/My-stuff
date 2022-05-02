""" I want to write a program that will automate peer reviewing files in Excel
looping, tab color checking, empty red box checking, Peer reivew: location """

######################################  Random code to use ######################################
from openpyxl import Workbook
from openpyxl import load_workbook
import os

ws.sheet_properties.tabColor.rgb #gives rgb value of tab color

""" unfilled color rgb = 255,128,128 """  

#################################################################################################


""" looping through folder """
# assign directory
directory = """ Assign folder here """  !!!!!!!
 
# iterate over files in that directory
for filename in os.listdir(directory):
    f = os.path.join(directory, filename)
    if os.path.isfile(f):    #checking if it's a file
      workbook = load_workbook(filename=f)   #loads the file as a workbook object
      workbook.sheetnames
      for sheet in workbook.sheetnames:
        if worbook.sheet_properties.tabColor != "Green":  #checks if tab color of sheet does not equal green
          pass        #so it does nothing
        else:
          
          
